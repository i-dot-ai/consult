"""Generate mock Question Understanding Excel fixtures for tests.

Run with: uv run python tests/fixtures/setup_consultation/create_qu_fixtures.py

Produces two files:
  qu_format_a_numeric.xlsx  — CUR047-style: 3 preamble rows, numeric question numbers
  qu_format_b_instruction.xlsx — LGR-style: 3 preamble rows + instruction row, label question numbers
"""

from pathlib import Path
import openpyxl

OUT = Path(__file__).parent


def _write_sheet(ws, rows):
    for row in rows:
        ws.append(row)


def make_format_a():
    """CUR047-style: no instruction row, numeric question numbers."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Demographic"
    _write_sheet(
        ws,
        [
            ["Demographic data", None],
            [
                "Must have a defined list of responses. Used to filter ALL questions.",
                None,
            ],
            ["Column ID", "Demographic title"],
            ["A", "Responding as"],
            ["F", "Size of organisation"],
        ],
    )

    ws = wb.create_sheet("Open questions")
    _write_sheet(
        ws,
        [
            ["Open Questions", None, None],
            [
                "Questions with free text that you want themes generating for. Not attached to any multiple choice questions.",
                None,
                None,
            ],
            [
                "Column ID",
                "Number",
                "Title of the question as you want it to appear in the app",
            ],
            [
                "I",
                2,
                "What role would you and/or your organisation play in achieving these outcomes?",
            ],
            [
                "K",
                3,
                "What will be the most important factors to achieving the proposed outcomes?",
            ],
            [
                "M",
                4,
                "What are the most important barriers that could prevent the proposed outcomes from being met?",
            ],
        ],
    )

    ws = wb.create_sheet("Hybrid questions")
    _write_sheet(
        ws,
        [
            ["Hybrid Questions", None, None, None],
            [
                "When a multiple choice and open choice question are connected. We need to know which 2 questions are connected.",
                None,
                None,
                None,
            ],
            [
                "Column ID - Open Question part",
                "Number",
                "Title of the question as you want it to appear in the app",
                "Column ID - Multiple Choice part",
            ],
            [
                "H",
                1,
                "Where do each of the proposed outcomes sit on a scale from very useful to not useful at all?",
                "G",
            ],
            [
                "Q",
                6,
                "Have you experienced any challenges with providing information via government digital services when complying with regulatory requirements?",
                "P",
            ],
        ],
    )

    ws = wb.create_sheet("Multiple Choice")
    _write_sheet(
        ws,
        [
            ["Mulitiple Choice Questions", None, None],
            [
                "Questions with a defined list of responses. Not attached to any open questions.",
                None,
                None,
            ],
            [
                "Column ID",
                "Number",
                "Title of the question as you want it to appear in the app",
            ],
            # No data rows — CUR047 had no closed-only questions
        ],
    )

    wb.save(OUT / "qu_format_a_numeric.xlsx")
    print("Written: qu_format_a_numeric.xlsx")


def make_format_b():
    """LGR-style: instruction row after column headers, label-style question numbers."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Demographic"
    _write_sheet(
        ws,
        [
            ["Demographic data", None],
            [
                "Must have a defined list of responses. Used to filter ALL questions.",
                None,
            ],
            ["Column ID", "Demographic title"],
            [
                "Which column does the question appear in on the export spreadsheet?",
                "Very short title to appear in the dashboard",
            ],
            ["A", "Respondent Type"],
            ["B", "Location"],
            ["C", "Organisation Type"],
        ],
    )

    ws = wb.create_sheet("Multiple Choice")
    _write_sheet(
        ws,
        [
            ["Mulitiple Choice Questions", None, None, None],
            [
                "Questions with a defined list of responses. Not attached to any open questions.",
                None,
                None,
                None,
            ],
            [
                "Column ID",
                "Number",
                "Title of the question as you want it to appear in the app",
                "If these are too long:",
            ],
            [
                "Which column does the question appear in on the export spreadsheet?",
                "Number the questions how they appear in the Consultation",
                "Remove any extra text the download has created",
                "If these are too long:",
            ],
            [
                "D",
                "Prop 1 Q1 - To what extent do you agree or disagree that the proposal suggests councils based on sensible geographies?",
                "Prop 1 Q1 - To what extent do you agree or disagree that the proposal suggests councils based on sensible geographies?",
                "P1 Q1",
            ],
            [
                "E",
                "Prop 1 Q2 - To what extent do you agree or disagree that the proposed councils will be able to deliver the outcomes?",
                "Prop 1 Q2 - To what extent do you agree or disagree that the proposed councils will be able to deliver the outcomes?",
                "P1 Q2",
            ],
            [
                "F",
                "Prop 1 Q3 - To what extent do you agree or disagree that the proposed councils are the right size to be efficient?",
                "Prop 1 Q3 - To what extent do you agree or disagree that the proposed councils are the right size to be efficient?",
                "P1 Q3",
            ],
        ],
    )

    ws = wb.create_sheet("Hybrid questions")
    _write_sheet(
        ws,
        [
            ["Hybrid Questions", None, None, None],
            [
                "When a multiple choice and open choice question are connected. We need to know which 2 questions are connected.",
                None,
                None,
                None,
            ],
            [
                "Column ID - Open Question part",
                "Number",
                "Title of the question as you want it to appear in the app",
                "Column ID - Multiple Choice part",
            ],
            [
                "Which column does the question appear in on the export spreadsheet?",
                "Number the questions how they appear in the Consultation",
                "Remove any extra text the download has created",
                "To create a hybrid question you must link 2 questions - put the column of the other question here",
            ],
            [
                "AC",
                "Proposal 3 - Q9 and Q10",
                "Prop 3 Q9 - To what extent do you agree the proposal sets out a strong public services justification for boundary change?",
                "AB",
            ],
            [
                "AM",
                "Proposal 4 - Q9 and Q10",
                "Prop 4 Q9 - To what extent do you agree the proposal sets out a strong public services justification for boundary change?",
                "AL",
            ],
        ],
    )

    ws = wb.create_sheet("Open questions")
    _write_sheet(
        ws,
        [
            ["Open Questions", None, None],
            [
                "Questions with free text that you want themes generating for. Not attached to any multiple choice questions.",
                None,
                None,
            ],
            [
                "Column ID",
                "Number",
                "Title of the question as you want it to appear in the app",
            ],
            [
                "Which column does the question appear in on the export spreadsheet?",
                "Number the questions how they appear in the Consultation",
                "Remove any extra text the download has created",
            ],
            [
                "K",
                "Proposal 1 Q8",
                "Prop 1 Q8 - If you would like to, please use the free text box to explain the answers you have provided to questions 1-7.",
            ],
            [
                "S",
                "Proposal 2 Q8",
                "Prop 2 Q8 - If you would like to, please use the free text box to explain the answers you have provided to questions 1-7.",
            ],
            [
                "AA",
                "Proposal 3 Q8",
                "Prop 3 Q8 - If you would like to, please use the free text box to explain the answers you have provided to questions 1-7.",
            ],
        ],
    )

    wb.save(OUT / "qu_format_b_instruction.xlsx")
    print("Written: qu_format_b_instruction.xlsx")


if __name__ == "__main__":
    make_format_a()
    make_format_b()
