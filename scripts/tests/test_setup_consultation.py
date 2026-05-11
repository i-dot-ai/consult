"""Tests for setup_consultation.py data loading and validation.

Fixture files in tests/fixtures/setup_consultation/:
  responses_lgr_format.csv       - single header row, immediate data (LGR Derbyshire-style)
  responses_scr_format.csv       - single header row + blank separator row (SCR CUR047-style)
  responses_biomass_format.csv   - two-tier header (short IDs row 0, full question text row 1)
  qu_format_a_numeric.xlsx       - CUR047-style: 3 preamble rows, numeric question numbers
  qu_format_b_instruction.xlsx   - LGR-style: 3 preamble rows + instruction row, label numbers
"""

import json
import pytest
import pandas as pd
import openpyxl
from pathlib import Path

from setup_consultation import (
    _load_qu_sheet,
    _normalise_null_like,
    _strip_control_chars,
    create_question_inputs,
    create_respondents_jsonl,
    load_and_number_question_sheets,
    load_responses,
    validate_data,
)

FIXTURES = Path(__file__).parent / "fixtures" / "setup_consultation"
QU_A = FIXTURES / "qu_format_a_numeric.xlsx"  # CUR047-style, numeric numbers
QU_B = (
    FIXTURES / "qu_format_b_instruction.xlsx"
)  # LGR-style, label numbers + instruction row


# ── load_responses ────────────────────────────────────────────────────────────


def test_load_responses_lgr_single_header():
    """LGR format: single header row, data immediately follows."""
    df, headers = load_responses(FIXTURES / "responses_lgr_format.csv")
    assert len(df) == 5
    assert "themefinder_id" in df.columns
    assert list(df["themefinder_id"]) == list(range(1, 6))
    assert headers["A"] == "Respondent type - Respondent type"
    assert (
        headers["F"]
        == "Prop 1 Q8 - Please use this space to explain your answers to questions 1-7"
    )


def test_load_responses_scr_blank_separator_dropped():
    """SCR format: blank separator row after header is dropped; data row count unaffected."""
    df, headers = load_responses(FIXTURES / "responses_scr_format.csv")
    assert len(df) == 5
    assert headers["A"] == "Responding as"
    # Blank row must not appear as a data row
    assert df["A"].isna().sum() == 0 or "nan" not in df["A"].astype(str).tolist()


def test_load_responses_biomass_two_tier_header():
    """Biomass format: row 0 short IDs discarded, row 1 long question text becomes the header."""
    df, headers = load_responses(FIXTURES / "responses_biomass_format.csv")
    assert len(df) == 5
    # Headers should come from the long-question-text row, not the short-ID row
    assert "Ident 1" not in headers.values()
    assert headers["A"] == "What is your name?"
    assert "organisation" in headers["C"].lower()


# ── _load_qu_sheet ─────────────────────────────────────────────────────────────


def test_load_qu_sheet_format_a_no_instruction_row():
    """Format A (CUR047): data starts immediately after 3 preamble rows; no instruction row to skip."""
    df = _load_qu_sheet(
        QU_A, "Open questions", 3, ["column_name", "question_number", "question_text"]
    )
    assert df is not None
    assert len(df) == 3
    assert df.iloc[0]["column_name"] == "I"
    assert df.iloc[0]["question_number"] == 2


def test_load_qu_sheet_format_b_skips_instruction_row():
    """Format B (LGR): instruction row after 3 preamble rows is detected and skipped."""
    df = _load_qu_sheet(
        QU_B, "Open questions", 3, ["column_name", "question_number", "question_text"]
    )
    assert df is not None
    assert len(df) == 3
    # First data row should be column K, not the instruction text
    assert df.iloc[0]["column_name"] == "K"
    assert "Which column" not in str(df.iloc[0]["question_number"])


def test_load_qu_sheet_demographic_format_a():
    """Demographic sheet loads cleanly in Format A with 2-column structure."""
    df = _load_qu_sheet(QU_A, "Demographic", 2, ["column_id", "label"])
    assert df is not None
    assert list(df["column_id"]) == ["A", "F"]
    assert list(df["label"]) == ["Responding as", "Size of organisation"]


def test_load_qu_sheet_demographic_format_b():
    """Demographic sheet loads cleanly in Format B, instruction row skipped."""
    df = _load_qu_sheet(QU_B, "Demographic", 2, ["column_id", "label"])
    assert df is not None
    assert list(df["column_id"]) == ["A", "B", "C"]
    assert "Which column" not in df["column_id"].tolist()


def test_load_qu_sheet_truncates_extra_columns():
    """LGR Multiple Choice has a 4th short-title column; _load_qu_sheet truncates to n_columns=3."""
    df = _load_qu_sheet(
        QU_B, "Multiple Choice", 3, ["column_name", "question_number", "question_text"]
    )
    assert df is not None
    assert list(df.columns) == ["column_name", "question_number", "question_text"]
    assert len(df) == 3
    assert df.iloc[0]["column_name"] == "D"


def test_load_qu_sheet_returns_none_for_missing_sheet():
    """Requesting a sheet that doesn't exist returns None without raising."""
    df = _load_qu_sheet(QU_A, "Nonexistent Sheet", 3, ["a", "b", "c"])
    assert df is None


def test_load_qu_sheet_strips_whitespace_from_column_ids(tmp_path):
    """Column IDs with trailing whitespace (e.g. 'C ') are stripped to 'C'."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Open questions"
    for row in [
        ["Open Questions", None, None],
        ["Description", None, None],
        ["Column ID", "Number", "Title"],
        ["C ", 1, "A question with a space in the column ID"],
        ["D", 2, "A question with a clean column ID"],
    ]:
        ws.append(row)
    path = tmp_path / "qu_spaces.xlsx"
    wb.save(path)

    df = _load_qu_sheet(
        path, "Open questions", 3, ["column_name", "question_number", "question_text"]
    )
    assert df is not None
    assert list(df["column_name"]) == ["C", "D"]


# ── load_and_number_question_sheets ──────────────────────────────────────────


def test_numbering_format_a_uses_numeric_question_numbers(capsys):
    """Format A: integer question numbers from the Q.U. file are used directly."""
    sheets = load_and_number_question_sheets(QU_A)
    assert "open" in sheets
    assert "hybrid" in sheets
    assert list(sheets["open"]["question_number"]) == [2, 3, 4]
    assert list(sheets["hybrid"]["question_number"]) == [1, 6]
    # Fallback warning should NOT appear
    out = capsys.readouterr().out
    assert "Non-numeric" not in out


def test_numbering_format_b_fallback_to_column_order(capsys):
    """Format B: label-style question numbers trigger column-order fallback numbering."""
    sheets = load_and_number_question_sheets(QU_B)
    out = capsys.readouterr().out
    assert "Non-numeric question numbers" in out

    # Numbers should be sequential integers assigned by Excel column order:
    # D(closed)=1, E(closed)=2, F(closed)=3, K(open)=4, S(open)=5, AA(open)=6, AC(hybrid)=7, AM(hybrid)=8
    assert list(sheets["closed"]["question_number"]) == [1, 2, 3]
    assert list(sheets["open"]["question_number"]) == [4, 5, 6]
    assert list(sheets["hybrid"]["question_number"]) == [7, 8]


def test_numbering_duplicate_question_numbers_raises(tmp_path):
    """Duplicate question numbers across sheets raise ValueError."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Open questions"
    for row in [
        ["Open Questions", None, None],
        ["Description", None, None],
        ["Column ID", "Number", "Title"],
        ["B", 1, "Open question 1"],
    ]:
        ws.append(row)

    ws2 = wb.create_sheet("Multiple Choice")
    for row in [
        ["Multiple Choice Questions", None, None],
        ["Description", None, None],
        ["Column ID", "Number", "Title"],
        ["C", 1, "Closed question also numbered 1"],  # duplicate!
    ]:
        ws2.append(row)

    # Required empty sheets
    for name in ["Hybrid questions", "Demographic"]:
        wb.create_sheet(name)

    path = tmp_path / "qu_dupes.xlsx"
    wb.save(path)

    with pytest.raises(ValueError, match="Non-unique question numbers"):
        load_and_number_question_sheets(path)


# ── validate_data ─────────────────────────────────────────────────────────────


def _responses(data: dict, n: int = 10) -> tuple[pd.DataFrame, dict[str, str]]:
    """Build a minimal responses_df and original_headers from column-name -> list pairs."""
    df = pd.DataFrame(data)
    df["themefinder_id"] = range(1, n + 1)
    headers = {chr(65 + i): col for i, col in enumerate(data.keys())}
    # Rename columns to Excel letters
    df.columns = [chr(65 + i) for i in range(len(data))] + ["themefinder_id"]
    return df, headers


def test_validate_data_clean(capsys):
    """Happy path: well-formed data produces no issues."""
    n = 30
    df, headers = _responses(
        {
            "Response ID": list(range(n)),
            "Please tell us your views on this proposal": [
                f"Unique response number {i} about the policy topic" for i in range(n)
            ],
            "Which option best describes your view": (
                ["Agree", "Disagree", "Strongly agree"] * (n // 3)
            ),
            "Region": ["North"] * (n // 2) + ["South"] * (n - n // 2),
        },
        n,
    )
    question_sheets = {
        "open": pd.DataFrame(
            {
                "column_name": ["B"],
                "question_number": [1],
                "question_text": ["Please tell us your views on this proposal"],
            }
        ),
        "closed": pd.DataFrame(
            {
                "column_name": ["C"],
                "question_number": [2],
                "question_text": ["Which option best describes your view"],
            }
        ),
    }
    validate_data(
        question_sheets,
        headers,
        df,
        demographic_columns=["D"],
        demographic_labels=["Region"],
        interactive=False,
    )
    out = capsys.readouterr().out
    assert "✓ Validation passed" in out


def test_validate_data_missing_column(capsys):
    """Q.U. references a column that does not exist in the response data."""
    n = 5
    df, headers = _responses(
        {
            "Response ID": list(range(n)),
            "Open text question": [f"Response {i}" for i in range(n)],
        },
        n,
    )
    question_sheets = {
        "open": pd.DataFrame(
            {
                "column_name": ["Z"],  # Z does not exist
                "question_number": [1],
                "question_text": ["Some open question"],
            }
        ),
    }
    validate_data(question_sheets, headers, df, interactive=False)
    out = capsys.readouterr().out
    assert "Z" in out
    assert "issue" in out.lower()


def test_validate_data_duplicate_column_reference(capsys):
    """Same column ID referenced in both Open questions and Multiple Choice sheets."""
    n = 10
    df, headers = _responses(
        {
            "Response ID": list(range(n)),
            "Question column": [
                f"Unique answer {i} about this topic in detail" for i in range(n)
            ],
        },
        n,
    )
    question_sheets = {
        "open": pd.DataFrame(
            {
                "column_name": ["B"],
                "question_number": [1],
                "question_text": ["Question column"],
            }
        ),
        "closed": pd.DataFrame(
            {
                "column_name": ["B"],  # B already used in open
                "question_number": [2],
                "question_text": ["Question column"],
            }
        ),
    }
    validate_data(question_sheets, headers, df, interactive=False)
    out = capsys.readouterr().out
    assert "more than once" in out


def test_validate_data_label_mismatch(capsys):
    """Q.U. question text has <40% string similarity with the response column header."""
    n = 5
    df, headers = _responses(
        {
            "Response ID": list(range(n)),
            "Name of the respondent organisation": [
                f"Some organisation {i}" for i in range(n)
            ],
        },
        n,
    )
    question_sheets = {
        "open": pd.DataFrame(
            {
                "column_name": ["B"],
                "question_number": [1],
                "question_text": [
                    "Do you support this planning proposal?"
                ],  # unrelated to header
            }
        ),
    }
    validate_data(question_sheets, headers, df, interactive=False)
    out = capsys.readouterr().out
    assert "low similarity" in out


def test_validate_data_number_mismatch(capsys):
    """Q.U. label and response header share similar text but contain different question numbers."""
    n = 5
    df, headers = _responses(
        {
            "Response ID": list(range(n)),
            "Question 7 - Tell us your views on this proposal": [
                f"Response {i}" for i in range(n)
            ],
        },
        n,
    )
    question_sheets = {
        "open": pd.DataFrame(
            {
                "column_name": ["B"],
                "question_number": [1],
                "question_text": ["Question 5 - Tell us your views on this proposal"],
            }
        ),
    }
    validate_data(question_sheets, headers, df, interactive=False)
    out = capsys.readouterr().out
    assert "number mismatch" in out


def test_validate_data_closed_looks_like_free_text(capsys):
    """A Multiple Choice column where all responses are unique triggers a free-text warning."""
    n = 20
    df, headers = _responses(
        {
            "Response ID": list(range(n)),
            "Which option": [
                f"Unique long response that reads like free text number {i}"
                for i in range(n)
            ],
        },
        n,
    )
    question_sheets = {
        "closed": pd.DataFrame(
            {
                "column_name": ["B"],
                "question_number": [1],
                "question_text": ["Which option"],
            }
        ),
    }
    validate_data(question_sheets, headers, df, interactive=False)
    out = capsys.readouterr().out
    assert "looks like free text" in out


def test_validate_data_open_looks_like_multichoice(capsys):
    """An Open questions column where nearly all responses are identical triggers a warning."""
    n = 20
    df, headers = _responses(
        {
            "Response ID": list(range(n)),
            "Tell us your views": ["Agree"] * 18
            + ["Disagree"] * 2,  # only 2 unique values
        },
        n,
    )
    question_sheets = {
        "open": pd.DataFrame(
            {
                "column_name": ["B"],
                "question_number": [1],
                "question_text": ["Tell us your views"],
            }
        ),
    }
    validate_data(question_sheets, headers, df, interactive=False)
    out = capsys.readouterr().out
    assert "Multiple Choice" in out


def test_validate_data_unreferenced_column(capsys):
    """Response columns not mentioned in any Q.U. sheet or demographics are flagged."""
    n = 5
    df, headers = _responses(
        {
            "Response ID": list(range(n)),
            "Open question": [f"Some response {i}" for i in range(n)],
            "Unlabelled column": [
                f"Data {i}" for i in range(n)
            ],  # not referenced anywhere
        },
        n,
    )
    question_sheets = {
        "open": pd.DataFrame(
            {
                "column_name": ["B"],
                "question_number": [1],
                "question_text": ["Open question"],
            }
        ),
    }
    validate_data(question_sheets, headers, df, interactive=False)
    out = capsys.readouterr().out
    assert "not referenced" in out
    assert "C" in out  # column C (Unlabelled column) should be named


# ── Deduplication ─────────────────────────────────────────────────────────────


def test_create_respondents_jsonl_deduplicates(tmp_path):
    """Comma-separated demographic values with duplicates are deduplicated, order preserved."""
    df = pd.DataFrame(
        {
            "themefinder_id": [1, 2, 3],
            "D": ["North,North,South", "East,East", "West,East,West"],
        }
    )
    create_respondents_jsonl(df, ["D"], ["Region"], tmp_path)

    lines = (tmp_path / "respondents.jsonl").read_text().splitlines()
    rows = [json.loads(line) for line in lines]
    assert rows[0]["demographic_data"]["Region"] == ["North", "South"]
    assert rows[1]["demographic_data"]["Region"] == ["East"]
    assert rows[2]["demographic_data"]["Region"] == ["West", "East"]


def test_create_respondents_jsonl_strips_trailing_comma(tmp_path):
    """Demographic value 'North,' must not produce a phantom '' bucket."""
    df = pd.DataFrame(
        {"themefinder_id": [1, 2], "D": ["North,", "North,South,"]}
    )
    create_respondents_jsonl(df, ["D"], ["Region"], tmp_path)
    rows = [
        json.loads(line)
        for line in (tmp_path / "respondents.jsonl").read_text().splitlines()
    ]
    assert rows[0]["demographic_data"]["Region"] == ["North"]
    assert rows[1]["demographic_data"]["Region"] == ["North", "South"]


def test_create_question_inputs_strips_trailing_comma_in_closed(tmp_path):
    """Closed question: 'Option A,' must not produce a phantom '' bucket."""
    df = pd.DataFrame(
        {
            "themefinder_id": [1, 2, 3, 4],
            "C": [
                "Option A,",            # single option with trailing comma
                "Option A,Option B,",   # two options with trailing comma
                "Option A,Option B,,",  # multiple trailing commas
                "Option A,Option B",    # no trailing comma — control case
            ],
        }
    )
    questions = [{"question_number": 1, "column_name": "C", "question_text": "?"}]
    create_question_inputs(df, questions, "closed", tmp_path)

    rows = [
        json.loads(line)
        for line in (
            tmp_path / "question_part_1" / "multi_choice.jsonl"
        ).read_text().splitlines()
    ]
    assert rows[0]["options"] == ["Option A"]
    assert rows[1]["options"] == ["Option A", "Option B"]
    assert rows[2]["options"] == ["Option A", "Option B"]
    assert rows[3]["options"] == ["Option A", "Option B"]

    q = json.loads((tmp_path / "question_part_1" / "question.json").read_text())
    assert "" not in q["multi_choice_options"]


def test_create_question_inputs_keeps_internal_empty_token_in_closed(tmp_path):
    """Internal empty tokens (e.g. 'A,,B') are preserved — only trailing strip."""
    df = pd.DataFrame(
        {"themefinder_id": [1], "C": ["A,,B"]}
    )
    questions = [{"question_number": 1, "column_name": "C", "question_text": "?"}]
    create_question_inputs(df, questions, "closed", tmp_path)
    rows = [
        json.loads(line)
        for line in (
            tmp_path / "question_part_1" / "multi_choice.jsonl"
        ).read_text().splitlines()
    ]
    assert rows[0]["options"] == ["A", "", "B"]


def test_create_question_inputs_strips_trailing_comma_in_hybrid(tmp_path):
    """Hybrid question: trailing comma on the closed side is also stripped."""
    df = pd.DataFrame(
        {
            "themefinder_id": [1],
            "B": ["Strongly agree,Agree,"],
            "C": ["Free-text answer"],
        }
    )
    questions = [
        {
            "question_number": 1,
            "open_column": "C",
            "closed_column": "B",
            "question_text": "?",
        }
    ]
    create_question_inputs(df, questions, "hybrid", tmp_path)
    rows = [
        json.loads(line)
        for line in (
            tmp_path / "question_part_1" / "multi_choice.jsonl"
        ).read_text().splitlines()
    ]
    assert rows[0]["options"] == ["Strongly agree", "Agree"]


def test_create_question_inputs_deduplicates_closed_options(tmp_path):
    """Closed question: comma-separated option values with duplicates are deduplicated."""
    df = pd.DataFrame(
        {
            "themefinder_id": [1, 2],
            "C": ["Option A,Option A,Option B", "Option B,Option B"],
        }
    )
    questions = [
        {"question_number": 1, "column_name": "C", "question_text": "Which option?"}
    ]
    create_question_inputs(df, questions, "closed", tmp_path)

    lines = (
        (tmp_path / "question_part_1" / "multi_choice.jsonl").read_text().splitlines()
    )
    rows = [json.loads(line) for line in lines]
    assert rows[0]["options"] == ["Option A", "Option B"]
    assert rows[1]["options"] == ["Option B"]


def test_create_question_inputs_deduplicates_hybrid_options(tmp_path):
    """Hybrid question: closed-part option values with duplicates are deduplicated."""
    df = pd.DataFrame(
        {
            "themefinder_id": [1, 2],
            "B": ["Strongly agree,Strongly agree,Agree", "Disagree,Disagree"],
            "C": [
                "I support this because it is well thought out",
                "I disagree due to the lack of evidence",
            ],
        }
    )
    questions = [
        {
            "question_number": 1,
            "open_column": "C",
            "closed_column": "B",
            "question_text": "Do you agree?",
        }
    ]
    create_question_inputs(df, questions, "hybrid", tmp_path)

    lines = (
        (tmp_path / "question_part_1" / "multi_choice.jsonl").read_text().splitlines()
    )
    rows = [json.loads(line) for line in lines]
    assert rows[0]["options"] == ["Strongly agree", "Agree"]
    assert rows[1]["options"] == ["Disagree"]


def test_create_question_inputs_strips_response_text_whitespace(tmp_path):
    """Open question: leading/trailing whitespace is stripped from response text."""
    df = pd.DataFrame(
        {
            "themefinder_id": [1, 2, 3],
            "C": [
                "  leading spaces",
                "trailing spaces   ",
                "\t\nwhitespace surrounded by newlines and tabs\n  ",
            ],
        }
    )
    questions = [{"question_number": 1, "column_name": "C", "question_text": "Why?"}]
    create_question_inputs(df, questions, "open", tmp_path)

    lines = (tmp_path / "question_part_1" / "responses.jsonl").read_text().splitlines()
    rows = [json.loads(line) for line in lines]
    assert rows[0]["text"] == "leading spaces"
    assert rows[1]["text"] == "trailing spaces"
    assert rows[2]["text"] == "whitespace surrounded by newlines and tabs"


# ── _normalise_null_like ──────────────────────────────────────────────────────


@pytest.mark.parametrize(
    "value",
    [
        "-",
        "",
        "   ",
        "  -  ",
    ],
)
def test_normalise_null_like_replaces_null_signifiers(value):
    df = pd.DataFrame({"themefinder_id": [1], "A": [value]})
    out = _normalise_null_like(df)
    assert out["A"].isna().all(), f"expected {value!r} to be treated as null"


@pytest.mark.parametrize(
    "value",
    [
        "A",
        "Yes",
        "No",
        "real answer",
        "Other",
        # Values intentionally NOT normalised — only "" and "-" are treated as
        # null-like. Anything else is preserved as a real answer.
        "--",
        "---",
        ".",
        "...",
        "N/A",
        "n/a",
        "n.a.",
        "NA",
        "None",
        "none",
        "NULL",
        "null",
        "nil",
        "nan",
        "Not Provided",
        "not applicable",
        "no response",
        "no answer",
        "no comment",
        "I have no comment to make on this",
        "A long free-text response that mentions N/A in passing",
        "0",
    ],
)
def test_normalise_null_like_preserves_real_answers(value):
    df = pd.DataFrame({"themefinder_id": [1], "A": [value]})
    out = _normalise_null_like(df)
    assert out["A"].iloc[0] == value


def test_normalise_null_like_leaves_themefinder_id_untouched():
    """themefinder_id is numeric and must never be normalised."""
    df = pd.DataFrame({"themefinder_id": [1, 2, 3], "A": ["-", "x", ""]})
    out = _normalise_null_like(df)
    assert list(out["themefinder_id"]) == [1, 2, 3]
    assert out["A"].isna().tolist() == [True, False, True]


def test_normalise_null_like_mixed_column():
    df = pd.DataFrame(
        {
            "themefinder_id": [1, 2, 3, 4, 5],
            "A": ["real", "-", "another", "", "  -  "],
        }
    )
    out = _normalise_null_like(df)
    assert out["A"].isna().tolist() == [False, True, False, True, True]
    assert out["A"].iloc[0] == "real"
    assert out["A"].iloc[2] == "another"


def test_load_responses_normalises_null_like_in_csv(tmp_path):
    """End-to-end: '-' cells in the CSV become NaN after loading."""
    csv = tmp_path / "responses.csv"
    csv.write_text(
        "Q1,Q2\n"
        "real answer,Agree\n"
        "-,Disagree\n"
        "another response,-\n"
    )
    df, _ = load_responses(csv)
    assert df["A"].isna().tolist() == [False, True, False]
    assert df["B"].isna().tolist() == [False, False, True]
    assert df["A"].iloc[0] == "real answer"


def test_create_question_inputs_skips_question_with_no_responses(tmp_path, capsys):
    """Closed question whose responses are all NaN: folder is not created."""
    df = pd.DataFrame(
        {
            "themefinder_id": [1, 2, 3],
            "A": [pd.NA, pd.NA, pd.NA],
        }
    )
    questions = [{"question_number": 7, "column_name": "A", "question_text": "?"}]
    create_question_inputs(df, questions, "closed", tmp_path)
    assert not (tmp_path / "question_part_7").exists()
    out = capsys.readouterr().out
    assert "Q7" in out and "skipping" in out


def test_create_question_inputs_skips_open_with_no_responses(tmp_path, capsys):
    df = pd.DataFrame({"themefinder_id": [1, 2], "C": [pd.NA, pd.NA]})
    questions = [{"question_number": 3, "column_name": "C", "question_text": "?"}]
    create_question_inputs(df, questions, "open", tmp_path)
    assert not (tmp_path / "question_part_3").exists()


def test_create_question_inputs_skips_hybrid_with_no_responses(tmp_path):
    df = pd.DataFrame(
        {
            "themefinder_id": [1, 2],
            "A": [pd.NA, pd.NA],
            "B": [pd.NA, pd.NA],
        }
    )
    questions = [
        {
            "question_number": 9,
            "open_column": "B",
            "closed_column": "A",
            "question_text": "?",
        }
    ]
    create_question_inputs(df, questions, "hybrid", tmp_path)
    assert not (tmp_path / "question_part_9").exists()


def test_create_question_inputs_still_writes_when_some_responses_present(tmp_path):
    """Sanity check: the skip only triggers when ALL responses are null."""
    df = pd.DataFrame(
        {"themefinder_id": [1, 2, 3], "A": ["Yes,No", pd.NA, "Maybe"]}
    )
    questions = [{"question_number": 4, "column_name": "A", "question_text": "?"}]
    create_question_inputs(df, questions, "closed", tmp_path)
    assert (tmp_path / "question_part_4" / "multi_choice.jsonl").exists()
    lines = (
        (tmp_path / "question_part_4" / "multi_choice.jsonl").read_text().splitlines()
    )
    assert len(lines) == 2


def test_load_responses_drops_rows_emptied_by_null_normalisation(tmp_path):
    """A row containing only null signifiers (e.g. all '-') is dropped entirely."""
    csv = tmp_path / "responses.csv"
    csv.write_text(
        "Q1,Q2\n"
        "real answer,Agree\n"
        "-,-\n"
        "another,Disagree\n"
    )
    df, _ = load_responses(csv)
    assert len(df) == 2
    assert df["A"].tolist() == ["real answer", "another"]


# ── _strip_control_chars / unicode preservation ──────────────────────────────


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("Community‑based", "Community‑based"),  # non-breaking hyphen kept
        ("en–dash", "en–dash"),
        ("em—dash", "em—dash"),
        ("smart ‘quote’", "smart ‘quote’"),
        ("smart “quote”", "smart “quote”"),
        ("ellipsis…", "ellipsis…"),
        ("café", "café"),
        ("£100", "£100"),
        # Control characters removed:
        ("bad\x00null", "badnull"),
        ("bell\x07char", "bellchar"),
        ("zero​width", "zerowidth"),
        ("﻿BOM", "BOM"),
    ],
)
def test_strip_control_chars(raw, expected):
    assert _strip_control_chars(raw) == expected


def test_clean_text_column_preserves_non_breaking_hyphen_in_options(tmp_path):
    """Q54-style: 'Community‑based' must round-trip through closed-question parsing."""
    df = pd.DataFrame(
        {"themefinder_id": [1, 2], "C": ["Community‑based", "Faith"]}
    )
    questions = [{"question_number": 1, "column_name": "C", "question_text": "?"}]
    create_question_inputs(df, questions, "closed", tmp_path)
    rows = [
        json.loads(line)
        for line in (
            tmp_path / "question_part_1" / "multi_choice.jsonl"
        ).read_text().splitlines()
    ]
    assert rows[0]["options"] == ["Community‑based"]
    assert rows[1]["options"] == ["Faith"]

    q = json.loads((tmp_path / "question_part_1" / "question.json").read_text())
    assert "Community‑based" in q["multi_choice_options"]
    assert "Communitybased" not in q["multi_choice_options"]


def test_create_respondents_jsonl_preserves_unicode(tmp_path):
    """Demographic values with accents / smart punctuation are not mangled."""
    df = pd.DataFrame(
        {"themefinder_id": [1], "D": ["café d‘Or"]}
    )
    create_respondents_jsonl(df, ["D"], ["Place"], tmp_path)
    row = json.loads((tmp_path / "respondents.jsonl").read_text().splitlines()[0])
    assert row["demographic_data"]["Place"] == ["café d‘Or"]
