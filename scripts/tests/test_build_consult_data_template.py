"""Smoke test for build_consult_data_template.py.

This script writes an xlsx with embedded formulas, dynamic-array array
formulas, and conditional formatting. The big risks are (a) the script
fails to run, (b) the workbook gets emitted but is structurally
corrupt, (c) one of the expected sheets is missing. This test catches
those by actually running the builder and re-opening the result with
openpyxl.

It doesn't validate the formulas themselves — those are exercised
manually by opening the workbook in Excel (the target deployment) and
visually checking the live audit. See the module docstring in
build_consult_data_template.py for the Excel-version compatibility
constraint the formulas have to satisfy.
"""

import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

SCRIPT = Path(__file__).resolve().parents[1] / "build_consult_data_template.py"


@pytest.fixture(scope="module")
def built_workbook(tmp_path_factory):
    """Build the template once per module and yield the .xlsx path."""
    out_dir = tmp_path_factory.mktemp("template-out")
    # OUTPUT_PATH is hard-coded next to the script; copy the script into
    # tmp_path so we don't clobber the committed example file.
    work = out_dir / "build_consult_data_template.py"
    work.write_bytes(SCRIPT.read_bytes())
    (out_dir / "build_consult_data_dummy_content.py").write_bytes(
        (SCRIPT.parent / "build_consult_data_dummy_content.py").read_bytes()
    )
    subprocess.run([sys.executable, str(work)], check=True, cwd=out_dir)
    produced = list(out_dir.glob("consult_data_template_*.xlsx"))
    assert len(produced) == 1, f"expected exactly one xlsx, got: {produced}"
    return produced[0]


def test_template_file_is_named_without_spaces(built_workbook):
    """Filename should be machine-friendly (no spaces, no ' - example ')."""
    assert " " not in built_workbook.name
    assert "example" not in built_workbook.name
    assert built_workbook.name.endswith(".xlsx")


def test_template_opens_cleanly(built_workbook):
    """openpyxl can re-open the file → it's not structurally corrupt."""
    wb = openpyxl.load_workbook(built_workbook)
    assert wb.sheetnames  # at least one sheet


def test_template_has_expected_sheets(built_workbook):
    """All sheets setup_consultation.py knows how to read must be present."""
    wb = openpyxl.load_workbook(built_workbook)
    names = set(wb.sheetnames)
    # Match the canonical names in QU_SHEET_SPECS + DEMOGRAPHIC_SHEET_ALIASES.
    required = {
        "Responses",
        "Demographics",
        "Open Questions",
        "Hybrid Questions",
        "Multiple Choice Questions",
    }
    missing = required - names
    assert not missing, f"missing sheets: {missing} (got: {sorted(names)})"
